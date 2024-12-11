import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
# Importar funciones desde ip2location.py
from IP2Location import (get_ip2location_info,calculate_time_difference,approximate_distance,is_impossible_travel,process_alarm)
# --- Función para obtener observaciones específicas de logs de Windows ---
def get_windows_login_observation(body):
    # Falta el Equipo (13/11)
    # Patrones para inglés y español
    patterns = [
        # Patrón para logs en inglés
        r'Alarm: Windows - Login.*?(\d+\.\d+\.\d+\.\d+).*?User: (.*?) Session ID:.*?Source Network Address: (\d+\.\d+\.\d+\.\d+)',
        # Patrón para logs en español
        r'Alarm: Windows - Login.*?(\d+\.\d+\.\d+\.\d+).*?Usuario: (.*?) Identificador de sesi.n:.*?Direcci.n de red de origen: (\d+\.\d+\.\d+\.\d+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, body, re.DOTALL)
        if match:
            ip = match.group(1).strip()
            user = match.group(2).strip()
            source_ip = match.group(3).strip()
            return f"Equipo: {ip} User: {user} Dirección de origen: {source_ip}"
    
    # Si no coincide con ninguno de los patrones
    return "No se pudo extraer información del inicio de sesión"

def get_linux_login_observation(body):
    pattern = r'Notificacion SIEM - Login sin usuario OPR o PS en Linux.*?Usuario: (.*?) Equipo: (.*?) Ip: (\d+\.\d+\.\d+\.\d+)'
    
    match = re.search(pattern, body, re.DOTALL)
    if match:
        user = match.group(1).strip()
        equipo = match.group(2).strip()
        ip = match.group(3).strip()
        return f"Usuario: {user} Equipo: {equipo} IP: {ip}"
    
    # Si no coincide el patrón
    return "No se pudo extraer información del login en Linux"



def get_sudo_su_observation(body):
    # Explicación del patrón:
    # - 'usuario: (.*?)' --> Captura cualquier usuario que sigue a la palabra 'usuario: '
    # - 'Cambio a: root' --> Esto es constante, no necesitamos capturarlo
    # - 'Host: (.*?)' --> Captura cualquier valor que sigue a 'Host: ' y se detiene en el espacio
    # - 'Ip: (\d+\.\d+\.\d+\.\d+)' --> Captura la dirección IP (formato de 4 bloques de números separados por puntos)
    
    pattern = r'usuario: (.*?) Cambio a: root Host: (.*?) Ip: (\d+\.\d+\.\d+\.\d+)'
    
    # Realizamos la búsqueda en el cuerpo
    match = re.search(pattern, body, re.DOTALL)
    
    # Si encontramos una coincidencia, extraemos los valores
    if match:
        user = match.group(1).strip()  # Captura del usuario
        host = match.group(2).strip()  # Captura del host
        ip = match.group(3).strip()    # Captura de la IP
        # Devolvemos la observación formateada
        return f"Usuario: {user} Host: {host} Ip: {ip}"
    
    # Si no hay coincidencia, devolvemos un mensaje por defecto
    return "No se pudo extraer información del sudo su"


def get_login_fuera_de_puentes_observation(body):
    # Primer patrón para "Login fuera de puentes"
    pattern_1 = r'Login fuera de puentes.*?Fecha/hora:.*?Usuario de origen: (.*?) IP de Origen: (\d+\.\d+\.\d+\.\d+) Host de Destino: (.*?) IPAM:'
    
    match_1 = re.search(pattern_1, body, re.DOTALL)
    if match_1:
        user = match_1.group(1).strip()
        ip_origen = match_1.group(2).strip()
        host_destino = match_1.group(3).strip()
        return f"Usuario: {user} IP de Origen: {ip_origen} Host de Destino: {host_destino} IPAM"
    
    # Segundo patrón para "Login sin usuario OPR o PS en Linux"
    pattern_2 = r'login en los sistemas linux.*?Usuario: (.*?) Equipo: (.*?) Ip: (\d+\.\d+\.\d+\.\d+)'
    
    match_2 = re.search(pattern_2, body, re.DOTALL)
    if match_2:
        user = match_2.group(1).strip()
        equipo = match_2.group(2).strip()
        ip = match_2.group(3).strip()
        return f"Usuario: {user} Equipo: {equipo} IP: {ip}"
    
    # Si ninguno de los patrones coincide
    return "No se pudo extraer información del login fuera de puentes"

def get_linux_scan_observation(body):
    try:
        # Dividir el texto en líneas y buscar "IP de Origen"
        lines = body.split()
        for i, word in enumerate(lines):
            if word == "Origen:":
                ip_origen = lines[i + 1]  # Obtener la IP que sigue a "IP de Origen:"
                
                # Verificar si la IP coincide con la específica
                if ip_origen == "10.150.49.89":
                    return "Escaneo de puertos detectado en Linux. Origen: Password Safe."
        
        # Si no se encuentra coincidencia, devolver None
        return None
    except Exception as e:
        # Manejo de errores
        print(f"Error al analizar el cuerpo del log: {e}")
        return None

# --- Código de normalización de logs ---
def normalize_body(body):
    patterns = [
        (r'^.*?(?=Usuario de origen)', 'Usuario de origen'),  # Patrón 1
        (r'^.*?(?=Usuario:)', 'Usuario:'),  # Patrón 2
        (r'^.*?(?=Se han detectado)', 'Se han detectado'),  # Patrón 3
        (r'Alarm: Windows - Login.*?(\d+\.\d+\.\d+\.\d+).*?\|\|3\|\|(.*?\.py).*?User: (.*?) Session ID:', 'Alarm: Windows - Login')  # Patrón 4 mejorado
    ]
    
    for pattern, start_phrase in patterns:
        if start_phrase in body:
            if start_phrase == 'Alarm: Windows - Login':
                # Manejo especial para el patrón de alarma de Windows
                match = re.search(pattern, body, re.DOTALL)
                if match:
                    ip = match.group(1).strip()
                    host = match.group(2).strip()
                    user = match.group(3).strip()
                    return f"Equipo {ip} Host: {host} User: {user}"
            else:
                normalized = re.sub(pattern, '', body, flags=re.DOTALL)
                return normalized.strip()
    
    # Si ningún patrón coincide, devolver el cuerpo original
    return body.strip()

def normalize_database(input_file):
    # Leer el archivo Excel
    df = pd.read_excel(input_file, sheet_name='BD_Logs')
    
    # Aplicar la normalización a la columna 'Cuerpo' y crear una nueva columna 'Cuerpo Normalizado'
    df['Cuerpo Normalizado'] = df['Cuerpo'].apply(normalize_body)
    
    # Devolver el DataFrame normalizado
    return df

# --- Código para manejar estilos ---
def apply_styles(workbook, sheet_name, row_index, col_index, is_critical=False, is_bold=False):
    sheet = workbook[sheet_name]
    cell = sheet.cell(row=row_index + 2, column=col_index + 1)
    
    if is_critical:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif is_bold:
        cell.font = Font(bold=True)
    else:
        cell.font = Font(bold=False)
        cell.fill = PatternFill(fill_type=None)

# --- Función para detectar alarmas críticas ---
def is_critical_alarm(alarm):
    critical_alarms = [
        "Notificacion SIEM - SIEM - Posible escaneo mediante VPN"
        # Añadir más alarmas críticas aquí
    ]
    return alarm in critical_alarms

def process_alarms(input_file, bd_file):
    # Leer y normalizar la base de datos
    df_bd = normalize_database(bd_file)
    
    # Leer el archivo de entrada (hoja '7-11')
    df_input = pd.read_excel(input_file, sheet_name='7-11')
    
    # Cargar el workbook
    workbook = load_workbook(input_file)
    sheet_input = workbook['7-11']  # Trabajamos en la hoja '7-11'
    
    # Procesar cada alarma en el archivo de entrada
    for index, row in df_input.iterrows():
        alarma = row['Alarma']
        cuerpo = row['Cuerpo']
        
        # Caso 1: Windows Login
        if alarma == "Notificacion SIEM - Se ha detectado un inicio de sesión":
            observacion = get_windows_login_observation(cuerpo)
            is_bold = True if observacion else False
        
        # Caso 2: Login sin usuario OPR o PS en Linux
        elif alarma == "Notificacion SIEM - Notificacion SIEM - Login sin usuario OPR o PS en Linux":
            observacion = get_login_fuera_de_puentes_observation(cuerpo)
            is_bold = True if observacion else False
        
        # Caso 3: Login fuera de puentes
        elif alarma == "Notificacion SIEM - Login fuera de puentes":
            observacion = get_login_fuera_de_puentes_observation(cuerpo)
            is_bold = True if observacion else False

        # Caso 4: sudo su detectado
        elif alarma == "Notificacion SIEM - Sudo su detectado":
            observacion = get_sudo_su_observation(cuerpo)
            is_bold = True if observacion else False
        # Casos de IP2Location
        elif alarma in [
            "Notificacion SIEM - VPN - Login desde 2 IPs diferentes",
            "Notificacion SIEM - Workapp - Login desde 2 IPs diferentes",
            "Notificacion SIEM - Workapp - Login fuera de ARG y PY",
            "Notificacion SIEM - VPN fuera de ARG o PY."
        ]:
            observacion = process_alarm(cuerpo)
            is_bold = True if observacion else False
        elif alarma == "Notificacion SIEM - Linux - Posible escaneo de puertos_local":
            observacion = get_linux_scan_observation(cuerpo)
            is_bold = True if observacion else False

        else:
            # Normalizar el cuerpo del log para matchear con la base de datos
            cuerpo_normalizado = normalize_body(cuerpo)
            
            # Buscar coincidencias en la base de datos normalizada
            matches = df_bd[df_bd['Cuerpo Normalizado'] == cuerpo_normalizado]
            
            if len(matches) >= 2:
                # Dos o más coincidencias
                observacion = matches.iloc[0]['Observación']
                is_bold = True
            elif len(matches) == 1:
                # Una coincidencia exacta
                observacion = matches.iloc[0]['Observación']
                is_bold = False
            else:
                # Sin coincidencias
                observacion = "Caso por defecto - Añadir manualmente"
                is_bold = False
        
        # Actualizar la observación en la columna B y el cuerpo en la columna C
        sheet_input.cell(row=index + 2, column=2, value=observacion)  # Columna B: Observación
        sheet_input.cell(row=index + 2, column=3, value=cuerpo)       # Columna C: Cuerpo (log)
        
        # Aplicar estilos a la columna B (col_index=1 porque es la segunda columna)
        is_critical = is_critical_alarm(alarma)
        apply_styles(workbook, '7-11', index, 1, is_critical, is_bold)
    
    # Guardar el resultado
    today = datetime.now().strftime("%d-%m")
    output_file = f"{today}_SIEM_DIA.xlsx"
    workbook.save(output_file)
    
    print(f"Procesamiento completado. Resultado guardado en {output_file}")



    

# Uso del script
input_file = "7-11.xlsx"
bd_file = "BD_Logs.xlsx"
process_alarms(input_file, bd_file)
