import pandas as pd
from openpyxl.styles import Font, PatternFill

def read_excel(file_path, sheet_name):
    """
    Lee un archivo Excel y devuelve un DataFrame para la hoja especificada.

    Args:
        file_path (str): Ruta al archivo Excel.
        sheet_name (str): Nombre de la hoja a leer.

    Returns:
        DataFrame: Contenido de la hoja en forma de DataFrame.
    """
    return pd.read_excel(file_path, sheet_name=sheet_name)

def save_excel(workbook, output_file):
    """
    Guarda el archivo Excel modificado.

    Args:
        workbook (Workbook): Objeto Workbook de openpyxl.
        output_file (str): Nombre del archivo de salida.
    """
    workbook.save(output_file)
    print(f"Archivo guardado como: {output_file}")

def apply_styles(sheet, row_index, is_critical=False, is_bold=False):
    """
    Aplica estilos de celda en una hoja de Excel.

    Args:
        sheet (Worksheet): Hoja de Excel donde aplicar estilos.
        row_index (int): Índice de la fila a modificar.
        is_critical (bool): Indica si la celda es crítica (rojo).
        is_bold (bool): Indica si el texto debe estar en negrita.
    """
    cell = sheet.cell(row=row_index + 2, column=2)  # Columna fija (índice 2)
    if is_critical:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif is_bold:
        cell.font = Font(bold=True)

def is_critical_alarm(alarm):
    """
    Verifica si una alarma es crítica.

    Args:
        alarm (str): Texto de la alarma.

    Returns:
        bool: True si la alarma es crítica, False en caso contrario.
    """
    critical_alarms = [
        "Notificacion SIEM - SIEM - Posible escaneo mediante VPN",
        # Agregar más alarmas críticas aquí
    ]
    return alarm in critical_alarms
