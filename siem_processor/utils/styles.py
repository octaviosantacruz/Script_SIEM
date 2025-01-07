from openpyxl.styles import Font, PatternFill

def apply_styles(sheet, row_index, is_critical=False, is_bold=False):
    cell = sheet.cell(row=row_index + 2, column=2)  # Adjust for 1-based indexing
    if is_critical:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    elif is_bold:
        cell.font = Font(bold=True)

def is_critical_alarm(alarm):
    critical_alarms = ["Notificacion SIEM - SIEM - Posible escaneo mediante VPN"]
    return alarm in critical_alarms
