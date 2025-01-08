"""
main.py | Este script procesa alarmas SIEM desde un archivo Excel y actualiza las observaciones basadas en la base de datos.
"""
import argparse
from openpyxl import load_workbook
from datetime import datetime
from siem_processor.utils.normalization import normalize_database
from siem_processor.utils.styles import apply_styles, is_critical_alarm
from siem_processor.cases.windows_login import handle_windows_login
from siem_processor.cases.linux_login import handle_linux_login
from siem_processor.cases.general_cases import handle_general_case
from siem_processor.cases.other_cases import handle_abm_cases, handle_salto_lateral_dba, handle_pases_produccion
#from siem_processor.modules.InfoGetterSIEM import fetch_user_info, get_user_details,extract_user_id
from siem_processor.modules.IP2Location import process_alarm
import pandas as pd
import os

# --- Función para procesar alarmas ---
def process_alarms(input_file,bd_file):
    """
    Procesa alarmas desde un archivo Excel de entrada y actualiza observaciones basadas en la base de datos.

    Args:
        input_file (str): Ruta al archivo Excel de entrada.
        bd_file (str): Ruta al archivo Excel de la base de datos.
    """
    # Normalizar la base de datos
    df_bd = normalize_database(bd_file)

    # Leer el archivo Excel de entrada
    df_input = pd.read_excel(input_file, sheet_name='7-11')
    workbook = load_workbook(input_file)
    sheet_input = workbook['7-11']

    for index, row in df_input.iterrows():
        alarma = row['Alarma']
        cuerpo = row['Cuerpo']

        # Procesar casos basados en el tipo de alarma
        if alarma == "Notificacion SIEM - Se ha detectado un inicio de sesión":
            observacion, is_bold = handle_windows_login(alarma, cuerpo)
        elif alarma in [
            "Notificacion SIEM - Login fuera de puentes",
            "Notificacion SIEM - Notificacion SIEM - Login sin usuario OPR o PS en Linux"
            "Notificacion SIEM - Sudo su detectado"
        ]:
            observacion, is_bold = handle_linux_login(alarma, cuerpo)
        elif alarma in [
            "Notificacion SIEM - VPN - Login desde 2 IPs diferentes",
            "Notificacion SIEM - Workapp - Login desde 2 IPs diferentes",
            "Notificacion SIEM - Workapp - Login fuera de ARG y PY",
            "Notificacion SIEM - VPN fuera de ARG o PY."
        ]:
            # Usar IP2Location para procesar alarmas de viajes imposibles
            observacion = process_alarm(cuerpo)
            is_bold = True if "Alerta" in observacion else False
        elif alarma in [
            "Notificacion SIEM - ABM-Usuario-AD-Creado",
            "Notificacion SIEM - ABM-Restablecimiento-Credenciales",
            "Notificacion SIEM - ABM-Grupo-AD-Agregado"

        ]:
            observacion, is_bold = handle_abm_cases(alarma, cuerpo)
            is_bold = True if "Alerta" in observacion else False
        elif alarma in [
            "Notificacion SIEM - Posible salto lateral 12+",
            "Notificacion SIEM - Posible salto lateral 6+"
        ]:
            observacion, is_bold = handle_salto_lateral_dba(alarma, cuerpo)
            is_bold = True if "Alerta" in observacion else False
        elif alarma in [
            "Notificacion SIEM - Notificacion SIEM - Pase a produccion detectado"
        ]:
            observacion, is_bold = handle_pases_produccion(alarma, cuerpo)
            is_bold = True if "Alerta" in observacion else False
        else:
            observacion, is_bold = handle_general_case(alarma, cuerpo)

        # Actualizar las observaciones en el Excel
        sheet_input.cell(row=index + 2, column=2, value=observacion)
        sheet_input.cell(row=index + 2, column=3, value=cuerpo)

        # Aplicar estilos para alarmas críticas y en negrita
        is_critical = is_critical_alarm(alarma)
        apply_styles(sheet_input, index, is_critical, is_bold)

    # Guardar el archivo procesado
    today = datetime.now().strftime("%d-%m")
    # Si no existe la carpeta output la crea
    if not os.path.exists("./siem_processor/output"):
        os.makedirs("./siem_processor/output")
    # Hacer cd a la carpeta output
    os.chdir("./siem_processor/output")
    output_file = f"{today}_SIEM_DIA.xlsx"
    workbook.save(output_file)
    print(f"Procesamiento completado. Resultado guardado en {output_file}")


# --- Punto de entrada principal ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Procesa alarmas SIEM desde un archivo Excel.")
    parser.add_argument(
        "--input_file", default="data/7-11.xlsx",
        help="Ruta al archivo Excel de entrada (por defecto: '7-11.xlsx')."
    )
    parser.add_argument(
        "--bd_file", default="data/BD_Logs.xlsx",
        help="Ruta al archivo Excel de la base de datos (por defecto: 'BD_Logs.xlsx')."
    )
    args = parser.parse_args()
    input_file = "C:/Users/u997568/Documents/GitHub/Script_SIEM/siem_processor/data/7-11.xlsx"
    bd_file = "C:/Users/u997568/Documents/GitHub/Script_SIEM/siem_processor/data/BD_Logs.xlsx"
    process_alarms(input_file=input_file, bd_file=bd_file)

